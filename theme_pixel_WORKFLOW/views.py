from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.views import LoginView, PasswordChangeView, PasswordResetView, PasswordResetConfirmView
from theme_pixel.forms import RegistrationForm, UserLoginForm, UserPasswordResetForm, UserPasswordChangeForm, UserSetPasswordForm
from django.contrib.auth import logout
from .models import *
from  django.contrib import messages
from django.contrib.auth.decorators import login_required , user_passes_test
from django.template.loader import render_to_string
from django.shortcuts import render
import weasyprint
from django.core.files.base import ContentFile
import os
import weasyprint
from django.http import HttpResponse
from django.template.loader import render_to_string

from django.conf import settings
from django.template.loader import get_template
from django.http import HttpResponse
import imgkit
from openpyxl import load_workbook
import io
import aspose.words as aw

from django.http import JsonResponse
from django.template.loader import get_template
from django.conf import settings
from html2image import Html2Image


from io import BytesIO
from zipfile import *

from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings
from html2image import Html2Image
import zipfile
import tempfile
import os

from PIL import Image
import csv
import pandas  as pd
import os
import openpyxl

import weasyprint

# Create your views here.



# Pages
@login_required

def index(request):
 
  products = Item.objects.all()
  context = {'products':products}

  return render(request, 'pages/indexe.html',context)

@user_passes_test(lambda u: u.is_superuser)
def Details(request : HttpRequest, id)-> HttpResponse:
  product =  Item.objects.get(id=id)
  rating = Rating.objects.filter(item=product, user=request.user).first()
  product.user_rating = rating.rating if rating else 0
  
  return render(request, 'pages/Details.html',{'products':product})

def rate(request: HttpRequest, id: int, rating: int) -> HttpResponse:
    post = Item.objects.get(id=id)
    Rating.objects.filter(item=post, user=request.user).delete()
    post.rating_set.create(user=request.user, rating=rating)
    return index(request)



def addProduct(request):
    if request.method == "POST":
        form = Item(request.POST, request.FILES)
        prod = Item()
        prod.name = request.POST.get('name')
        prod.image_details=request.FILES['image_details']
        prod.description=request.POST.get('description')

        if len(request.FILES) != 0:
            prod.image = request.FILES['image']
        
        prod.file = request.FILES['file']
          
            
       

        prod.save()
        messages.success(request, "PLan Added Successfully")
        return redirect('/')
    return render(request, 'pages/add.html')
@login_required
def dashboard(request):
  
  return render(request, 'pages/dashboard.html')  

@login_required
def csv_view(request):
  
  image = Image.open('env/Lib/site-packages/theme_pixel/static/assets/img/options/OPTIONBI.jpg')
  pixels = list(image.getdata())
  width, height = image.size
  data = []
  for i in range(height):
    row = pixels[i * width : (i + 1) * width]
    row_str = ','.join(str(p) for p in row)
    data.append(row_str)
  csv_data = '\n'.join(data)
  with open('data.csv', 'w') as f:
    f.write(csv_data)
  with open('data.csv', newline='') as csvfile:
    reader = csv.reader(csvfile, delimiter=',', quotechar='"')
    data = [row for row in reader]
  with open('data.csv', newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', quotechar='"')
        data = [row for row in reader]
        context = {'data': data}
  return render(request, 'pages/csv_reader.html', context)



import openpyxl

# def datatable(request):
#     workbook = openpyxl.load_workbook(filename='env/Lib/site-packages/theme_pixel/static/assets/csv/Plan_Cloud.xlsx')
#     worksheet = workbook.active
#     headers = [cell.value for cell in worksheet[1]]
#     out = []
#     for row in worksheet.iter_rows(min_row=2):
#         row_data = tuple(cell.value for cell in row)
#         out.append(row_data)
#     return render(request, 'pages/datatable.html', {'data' : out, 'headers' : headers})

def datatable(request : HttpRequest, id)-> HttpResponse:
    excel_file = Item.objects.get(id=id)
    xlsx_data = excel_file.file
    workbook = openpyxl.load_workbook(filename=xlsx_data, data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    out = []
    for row in worksheet.iter_rows(min_row=2):
        row_data = tuple(cell.value for cell in row)
        out.append(row_data)
    return render(request, 'pages/datatable.html', {'data' : out, 'headers' : headers})

    
def excel_to_png(request: HttpRequest, id)-> HttpResponse:
    
    excel_file = Item.objects.get(id=id)
    xlsx_data = excel_file.file
    # Load the Excel file
    wb = openpyxl.load_workbook(filename=xlsx_data, data_only=True)
    sheet = wb.active

    # Get the dimensions of the worksheet
    width = sheet.max_column * 100
    height = sheet.max_row * 30

    # Create a new image with the same dimensions
    image = Image.new('RGB', (width, height), (255, 255, 255))
    pixels = image.load()

    # Loop through the worksheet and fill in the image pixels
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            cell_value = str(sheet.cell(row=i, column=j).value)
            pixels[j*100-100, i*30-30] = (0, 0, 0)  # set pixel color for cell border
            pixels[j*100-1, i*30-1] = (0, 0, 0)  # set pixel color for cell border
            pixels[j*100-50, i*30-15] = (0, 0, 0)  # set pixel color for cell value
            image.putpixel((j*100-50, i*30-15), (0, 0, 0))  # set pixel color for cell value
            # Put the cell value in the image
            # You may need to adjust the pixel positions based on the dimensions of the cells and the font size
            # For example, in the code above, we assume a cell is 100 pixels wide and 30 pixels tall, and we put the cell value in the center of the cell

    # Save the image as PNG
    buffer = BytesIO()
    image.save(buffer, format='PNG')
    buffer.seek(0)

    # Return the image as HTTP response
    response = HttpResponse(buffer, content_type='image/png')
    response['Content-Disposition'] = 'attachment; filename="excel.png"'
    return response

def FormulaireBI(request):
    return render(request, 'pages/formulaireBi.html')

def abouts_us(request):
  return render(request, 'pages/about.html')

def contact_us(request):
  return render(request, 'pages/contact-us.html')

def landing_freelancer(request):
  return render(request, 'pages/landing-freelancer.html')

def blank_page(request):
  return render(request, 'pages/blank.html')


# Authentication
class UserLoginView(LoginView):
  template_name = 'accounts/sign-in.html'
  form_class = UserLoginForm

def logout_view(request):
  logout(request)
  return redirect('/accounts/login/')

def register(request):
  if request.method == 'POST':
    form = RegistrationForm(request.POST)
    if form.is_valid():
      form.save()
      print("Account created successfully!")
      return redirect('/accounts/login')
    else:
      print("Registration failed!")
  else:
    form = RegistrationForm()

  context = { 'form': form }
  return render(request, 'accounts/sign-up.html', context)


class UserPasswordResetView(PasswordResetView):
  template_name = 'accounts/password_reset.html'
  form_class = UserPasswordResetForm

class UserPasswordResetConfirmView(PasswordResetConfirmView):
  template_name = 'accounts/password_reset_confirm.html'
  form_class = UserSetPasswordForm

class UserPasswordChangeView(PasswordChangeView):
  template_name = 'accounts/password_change.html'
  form_class = UserPasswordChangeForm

from PIL import Image, ImageDraw, ImageFont

def excel_to_image(request, id):
   
    # Retrieve the Excel file object using the ID parameter
    excel_file = Item.objects.get(id=id)
    xlsx_data = excel_file.file.path

    # Charger le fichier Excel en utilisant openpyxl
    wb = openpyxl.load_workbook(filename=xlsx_data, data_only=True)
    sheet = wb.active

    # Set up the image
    cell_width = 250
    cell_height = 50
    font_size = 20
    padding = 20
    margin = 20
    num_cols = sheet.max_column
    num_rows = sheet.max_row
    image_width = cell_width * num_cols
    image_height = cell_height * num_rows
    image = Image.new('RGB', (image_width, image_height), color='white')
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype('arial.ttf', font_size)

    # Draw the table
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        for j, cell in enumerate(row):
            x = j * cell_width
            y = i * cell_height
            draw.rectangle((x, y, x + cell_width, y + cell_height), outline='black')
            draw.text((x + padding, y + padding), str(cell), font=font, fill='black')

    # Save the image
    buffer = BytesIO()
    image.save(buffer, format='JPEG')
    buffer.seek(0)

    # Update the file field of the Item object with the new image
    excel_file.image_details.save('table.jpeg', ContentFile(buffer.getvalue()))
    response = HttpResponse(buffer, content_type='image/png')
    response['Content-Disposition'] = 'attachment; filename="table.png"'
    
    return redirect('index')
    # Return a redirect response to the original view that displays the updated image
  
